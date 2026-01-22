#!/usr/bin/env python3
"""
StudyHub Data Setup Script
==========================

This script helps you:
1. Create a properly formatted Google Sheet from the Excel template
2. Convert existing data to the required format
3. Validate your data structure
4. Generate sample data for testing

Usage:
    python setup_data.py --help
    python setup_data.py create-sample
    python setup_data.py validate path/to/data.xlsx
    python setup_data.py export-json path/to/data.xlsx
"""

import json
import os
import sys
from datetime import datetime

# Check for required packages
try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except ImportError:
    print("Installing required packages...")
    os.system(f"{sys.executable} -m pip install pandas openpyxl --break-system-packages")
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ============================================================================
# DATA STRUCTURE DEFINITIONS
# ============================================================================

SHEET_SCHEMAS = {
    'Subjects': {
        'columns': ['subject_id', 'subject_key', 'name', 'icon', 'color_hex', 'light_bg', 'gradient_from', 'gradient_to', 'dark_glow'],
        'required': ['subject_id', 'subject_key', 'name'],
        'description': 'Define subjects with their visual styling'
    },
    'Topics': {
        'columns': ['topic_id', 'subject_key', 'topic_name', 'duration_minutes', 'order_index'],
        'required': ['topic_id', 'subject_key', 'topic_name'],
        'description': 'List all topics per subject'
    },
    'Topic_Sections': {
        'columns': ['section_id', 'topic_id', 'section_title', 'section_icon', 'order_index', 'section_type'],
        'required': ['section_id', 'topic_id', 'section_title'],
        'description': 'Define sections/chapters within each topic'
    },
    'Learning_Objectives': {
        'columns': ['objective_id', 'topic_id', 'objective_text', 'order_index'],
        'required': ['objective_id', 'topic_id', 'objective_text'],
        'description': 'Learning objectives for each topic'
    },
    'Key_Terms': {
        'columns': ['term_id', 'topic_id', 'term', 'definition'],
        'required': ['term_id', 'topic_id', 'term', 'definition'],
        'description': 'Vocabulary terms and definitions'
    },
    'Study_Content': {
        'columns': ['content_id', 'section_id', 'content_type', 'content_title', 'content_text', 'order_index', 'image_url', 'video_url'],
        'required': ['content_id', 'section_id', 'content_type', 'content_text'],
        'description': 'Main educational content blocks'
    },
    'Formulas': {
        'columns': ['formula_id', 'topic_id', 'formula_text', 'formula_label', 
                   'variable_1_symbol', 'variable_1_name', 'variable_1_unit',
                   'variable_2_symbol', 'variable_2_name', 'variable_2_unit',
                   'variable_3_symbol', 'variable_3_name', 'variable_3_unit'],
        'required': ['formula_id', 'topic_id', 'formula_text'],
        'description': 'Mathematical/scientific formulas'
    },
    'Quiz_Questions': {
        'columns': ['question_id', 'topic_id', 'question_text', 'option_a', 'option_b', 'option_c', 'option_d', 
                   'correct_answer', 'explanation', 'xp_reward'],
        'required': ['question_id', 'topic_id', 'question_text', 'option_a', 'option_b', 'correct_answer'],
        'description': 'Multiple choice quiz questions'
    },
    'Achievements': {
        'columns': ['achievement_id', 'icon', 'name', 'description', 'unlock_condition'],
        'required': ['achievement_id', 'name', 'description'],
        'description': 'Gamification badges and achievements'
    }
}

CONTENT_TYPES = ['introduction', 'formula', 'concept_helper', 'warning', 'real_world', 'text', 'video', 'image', 'flowchart']
SECTION_TYPES = ['objectives', 'intro', 'content', 'applications', 'quiz']
VALID_ICONS = ['Zap', 'Calculator', 'FlaskConical', 'Leaf', 'Trophy', 'Star', 'Award', 'Flame',
               'HelpCircle', 'CheckCircle2', 'Target', 'BookOpen', 'FileText', 'Clock', 'Globe',
               'Lightbulb', 'AlertTriangle', 'Atom', 'Microscope', 'Dna', 'Pi', 'Hammer', 'RefreshCw',
               'Minimize2', 'Triangle', 'Disc', 'Grid', 'ArrowDown', 'Link', 'GitCommit', 'Circle',
               'GitBranch', 'Share2']


# ============================================================================
# SAMPLE DATA
# ============================================================================

SAMPLE_DATA = {
    'Subjects': [
        {'subject_id': 'phys-001', 'subject_key': 'physics', 'name': 'Physics', 'icon': 'Zap', 
         'color_hex': '#3B82F6', 'light_bg': 'bg-blue-50', 'gradient_from': 'blue-500', 
         'gradient_to': 'blue-600', 'dark_glow': 'shadow-blue-500/20'},
        {'subject_id': 'math-001', 'subject_key': 'math', 'name': 'Mathematics', 'icon': 'Calculator',
         'color_hex': '#10B981', 'light_bg': 'bg-emerald-50', 'gradient_from': 'emerald-500',
         'gradient_to': 'emerald-600', 'dark_glow': 'shadow-emerald-500/20'},
        {'subject_id': 'chem-001', 'subject_key': 'chemistry', 'name': 'Chemistry', 'icon': 'FlaskConical',
         'color_hex': '#F59E0B', 'light_bg': 'bg-amber-50', 'gradient_from': 'amber-500',
         'gradient_to': 'amber-600', 'dark_glow': 'shadow-amber-500/20'},
        {'subject_id': 'bio-001', 'subject_key': 'biology', 'name': 'Biology', 'icon': 'Leaf',
         'color_hex': '#8B5CF6', 'light_bg': 'bg-violet-50', 'gradient_from': 'violet-500',
         'gradient_to': 'violet-600', 'dark_glow': 'shadow-violet-500/20'},
    ],
    'Topics': [
        # PHYSICS
        {'topic_id': 'phys-t1', 'subject_key': 'physics', 'topic_name': "Newton's Laws", 'duration_minutes': 30, 'order_index': 1},
        {'topic_id': 'phys-t2', 'subject_key': 'physics', 'topic_name': 'Work & Energy', 'duration_minutes': 45, 'order_index': 2},
        {'topic_id': 'phys-t3', 'subject_key': 'physics', 'topic_name': 'Electricity', 'duration_minutes': 40, 'order_index': 3},

        # MATH
        {'topic_id': 'math-t1', 'subject_key': 'math', 'topic_name': 'Algebraic Expressions', 'duration_minutes': 35, 'order_index': 1},
        {'topic_id': 'math-t2', 'subject_key': 'math', 'topic_name': 'Geometry: Triangles', 'duration_minutes': 30, 'order_index': 2},
        {'topic_id': 'math-t3', 'subject_key': 'math', 'topic_name': 'Probability', 'duration_minutes': 25, 'order_index': 3},

        # CHEMISTRY
        {'topic_id': 'chem-t1', 'subject_key': 'chemistry', 'topic_name': 'Atomic Structure', 'duration_minutes': 40, 'order_index': 1},
        {'topic_id': 'chem-t2', 'subject_key': 'chemistry', 'topic_name': 'The Periodic Table', 'duration_minutes': 35, 'order_index': 2},
        {'topic_id': 'chem-t3', 'subject_key': 'chemistry', 'topic_name': 'Chemical Bonding', 'duration_minutes': 45, 'order_index': 3},

        # BIOLOGY
        {'topic_id': 'bio-t1', 'subject_key': 'biology', 'topic_name': 'Cell Biology', 'duration_minutes': 30, 'order_index': 1},
        {'topic_id': 'bio-t2', 'subject_key': 'biology', 'topic_name': 'Genetics & DNA', 'duration_minutes': 40, 'order_index': 2},
        {'topic_id': 'bio-t3', 'subject_key': 'biology', 'topic_name': 'Ecosystems', 'duration_minutes': 35, 'order_index': 3},
    ],
    'Topic_Sections': [],
    'Learning_Objectives': [],
    'Key_Terms': [],
    'Study_Content': [],
    'Formulas': [],
    'Quiz_Questions': [],
    'Achievements': [
        {'achievement_id': 'first-login', 'icon': 'Zap', 'name': 'First Login', 'description': 'Welcome to StudyHub!', 'unlock_condition': 'Login for the first time'},
        {'achievement_id': 'first-quiz', 'icon': 'HelpCircle', 'name': 'First Quiz', 'description': 'Complete your first quiz', 'unlock_condition': 'Complete any quiz'},
        {'achievement_id': 'streak-5', 'icon': 'Flame', 'name': '5-Day Streak', 'description': 'Study 5 days in a row', 'unlock_condition': 'streak >= 5'},
        {'achievement_id': 'streak-10', 'icon': 'Flame', 'name': '10-Day Streak', 'description': 'Study 10 days in a row', 'unlock_condition': 'streak >= 10'},
        {'achievement_id': 'topic-complete', 'icon': 'CheckCircle2', 'name': 'Topic Master', 'description': 'Complete any topic', 'unlock_condition': 'Any topic progress = 100'},
        {'achievement_id': 'subject-50', 'icon': 'Trophy', 'name': 'Halfway There', 'description': '50% in any subject', 'unlock_condition': 'Any subject progress >= 50'},
        {'achievement_id': 'perfect-quiz', 'icon': 'Star', 'name': 'Perfect Score', 'description': 'Score 100% on a quiz', 'unlock_condition': 'Any quiz score = 100'},
        {'achievement_id': 'all-subjects', 'icon': 'Award', 'name': 'Well Rounded', 'description': 'Study all 4 subjects', 'unlock_condition': 'All subjects accessed'},
    ]
}

# Helper to add a topic's data
def add_topic_data(topic_id, sections, objectives, terms, content, formulas, questions):
    # Add sections
    for i, section in enumerate(sections, 1):
        SAMPLE_DATA['Topic_Sections'].append({
            'section_id': f"{topic_id}-s{i}",
            'topic_id': topic_id,
            'section_title': section['title'],
            'section_icon': section.get('icon', 'FileText'),
            'order_index': i,
            'section_type': section.get('type', 'content')
        })

    # Add objectives
    for i, obj in enumerate(objectives, 1):
        SAMPLE_DATA['Learning_Objectives'].append({
            'objective_id': f"obj-{topic_id}-{i}",
            'topic_id': topic_id,
            'objective_text': obj,
            'order_index': i
        })

    # Add key terms
    for i, term in enumerate(terms, 1):
        SAMPLE_DATA['Key_Terms'].append({
            'term_id': f"term-{topic_id}-{i}",
            'topic_id': topic_id,
            'term': term['term'],
            'definition': term['def']
        })

    # Add content
    for c in content:
        SAMPLE_DATA['Study_Content'].append({
            'content_id': f"cont-{topic_id}-{len(SAMPLE_DATA['Study_Content'])+1}",
            'section_id': f"{topic_id}-s{c['sec_idx']}",
            'content_type': c['type'],
            'content_title': c.get('title', ''),
            'content_text': c['text'],
            'order_index': c.get('order', 1),
            'image_url': c.get('image_url', ''),
            'video_url': c.get('video_url', '')
        })

    # Add formulas
    for i, f in enumerate(formulas, 1):
        SAMPLE_DATA['Formulas'].append({
            'formula_id': f"form-{topic_id}-{i}",
            'topic_id': topic_id,
            'formula_text': f['text'],
            'formula_label': f['label'],
            'variable_1_symbol': f.get('v1s', ''), 'variable_1_name': f.get('v1n', ''), 'variable_1_unit': f.get('v1u', ''),
            'variable_2_symbol': f.get('v2s', ''), 'variable_2_name': f.get('v2n', ''), 'variable_2_unit': f.get('v2u', ''),
            'variable_3_symbol': f.get('v3s', ''), 'variable_3_name': f.get('v3n', ''), 'variable_3_unit': f.get('v3u', ''),
        })

    # Add quizzes
    for i, q in enumerate(questions, 1):
        SAMPLE_DATA['Quiz_Questions'].append({
            'question_id': f"quiz-{topic_id}-{i}",
            'topic_id': topic_id,
            'question_text': q['text'],
            'option_a': q['a'], 'option_b': q['b'], 'option_c': q['c'], 'option_d': q['d'],
            'correct_answer': q['ans'],
            'explanation': q['exp'],
            'xp_reward': 10
        })

# ==========================================
# POPULATE DETAILED CONTENT
# ==========================================

# 1. PHYSICS - Newton's Laws
add_topic_data('phys-t1',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Introduction', 'icon': 'BookOpen', 'type': 'intro'},
        {'title': 'First Law (Inertia)', 'icon': 'Zap', 'type': 'content'},
        {'title': 'Second Law (F=ma)', 'icon': 'Calculator', 'type': 'content'},
        {'title': 'Third Law (Action-Reaction)', 'icon': 'Zap', 'type': 'content'},
        {'title': 'Assessment', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Define inertia and its relationship to mass', 'Apply F=ma to solve problems', 'Identify action-reaction pairs', 'Understand the concept of net force', 'Distinguish between mass and weight'],
    terms=[
        {'term': 'Inertia', 'def': 'Resistance of any physical object to any change in its velocity'},
        {'term': 'Force', 'def': 'A push or pull upon an object resulting from interaction with another object'},
        {'term': 'Mass', 'def': 'A measure of the amount of matter in an object'},
        {'term': 'Net Force', 'def': 'The vector sum of all forces acting on an object'},
        {'term': 'Acceleration', 'def': 'The rate of change of velocity per unit of time'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'The Foundations of Dynamics', 'text': "Isaac Newton's three laws of motion describe the relationship between the motion of an object and the forces acting on it.", 'image_url': 'https://images.unsplash.com/photo-1451187580459-43490279c0fa?w=800'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Law of Inertia', 'text': 'An object at rest stays at rest and an object in motion stays in motion unless acted upon by an unbalanced force.'},
        {'sec_idx': 3, 'type': 'real_world', 'title': 'Seatbelts', 'text': 'When a car stops suddenly, your body keeps moving forward due to inertia. Seatbelts provide the unbalanced force to stop you.'},
        {'sec_idx': 3, 'type': 'warning', 'title': 'Inertia is NOT a force', 'text': 'Inertia is a property of matter, not a force that pushes you.'},
        {'sec_idx': 4, 'type': 'formula', 'title': 'The Equation', 'text': 'F = ma'},
        {'sec_idx': 4, 'type': 'text', 'title': 'Explanation', 'text': 'Force equals mass times acceleration. The more mass an object has, the more force is needed to accelerate it.'},
         {'sec_idx': 4, 'type': 'concept_helper', 'title': 'Proportionality', 'text': 'Acceleration is directly proportional to Force and inversely proportional to Mass.'},
        {'sec_idx': 5, 'type': 'text', 'title': 'Symmetry in Forces', 'text': 'For every action, there is an equal and opposite reaction. Forces always come in pairs.'},
        {'sec_idx': 5, 'type': 'warning', 'title': 'Common Mistake', 'text': 'Action and reaction forces act on DIFFERENT objects, so they do not cancel each other out!'},
        {'sec_idx': 5, 'type': 'real_world', 'title': 'Rocket Propulsion', 'text': 'A rocket pushes gas down (action), and the gas pushes the rocket up (reaction).'}
    ],
    formulas=[
        {'text': 'F = m \\cdot a', 'label': "Newton's Second Law", 'v1s': 'F', 'v1n': 'Force', 'v1u': 'N', 'v2s': 'm', 'v2n': 'Mass', 'v2u': 'kg', 'v3s': 'a', 'v3n': 'Acceleration', 'v3u': 'm/s²'},
        {'text': 'W = m \\cdot g', 'label': "Weight", 'v1s': 'W', 'v1n': 'Weight', 'v1u': 'N', 'v2s': 'm', 'v2n': 'Mass', 'v2u': 'kg', 'v3s': 'g', 'v3n': 'Gravity', 'v3u': 'm/s²'}
    ],
    questions=[
        {'text': 'Which property of an object determines its inertia?', 'a': 'Volume', 'b': 'Mass', 'c': 'Weight', 'd': 'Velocity', 'ans': 'B', 'exp': 'Mass is a direct measure of inertia.'},
        {'text': 'If you double the force on an object, what happens to its acceleration?', 'a': 'Doubles', 'b': 'Halves', 'c': 'Quadruples', 'd': 'Stays same', 'ans': 'A', 'exp': 'Acceleration is directly proportional to force (F=ma).'},
        {'text': 'A 10kg object accelerates at 2 m/s². What is the force?', 'a': '5 N', 'b': '12 N', 'c': '20 N', 'd': '0.2 N', 'ans': 'C', 'exp': 'F = m * a = 10 * 2 = 20 N.'},
        {'text': 'Which law explains why a book sits still on a table?', 'a': '1st Law', 'b': '2nd Law', 'c': '3rd Law', 'd': 'Gravitational Law', 'ans': 'A', 'exp': '1st Law: Objects at rest stay at rest unless acted on by unbalanced force.'},
        {'text': 'Action and reaction forces are always...', 'a': 'Unequal', 'b': 'In the same direction', 'c': 'Acting on the same object', 'd': 'Equal and opposite', 'ans': 'D', 'exp': 'Newton\'s 3rd Law states they are equal in magnitude and opposite in direction.'}
    ]
)

# 2. PHYSICS - Work & Energy
add_topic_data('phys-t2',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Work', 'icon': 'Hammer', 'type': 'content'},
        {'title': 'Energy Types', 'icon': 'Zap', 'type': 'content'},
        {'title': 'Conservation', 'icon': 'RefreshCw', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Define Work in physics', 'Distinguish between kinetic and potential energy', 'Apply conservation of energy principle', 'Calculate work and power', 'Understand mechanical advantage'],
    terms=[
        {'term': 'Work', 'def': 'Force applied over a distance (Joules)'},
        {'term': 'Kinetic Energy', 'def': 'Energy of motion'},
        {'term': 'Potential Energy', 'def': 'Stored energy due to position or state'},
        {'term': 'Power', 'def': 'The rate at which work is done (Watts)'},
        {'term': 'Mechanical Energy', 'def': 'Sum of potential and kinetic energy'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Physics Definition of Work', 'text': 'In physics, work is done only when a force moves an object. Pushing a wall and not moving it means zero work is done!', 'image_url': 'https://images.unsplash.com/photo-1516937941348-c096b542b9c9?w=800'},
        {'sec_idx': 2, 'type': 'formula', 'title': 'Work Formula', 'text': 'W = F \\cdot d'},
        {'sec_idx': 2, 'type': 'concept_helper', 'title': 'Direction Matters', 'text': 'The force must be in the same direction as the movement for maximum work.'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Kinetic vs Potential', 'text': 'A roller coaster at the top has high Potential Energy. As it falls, it converts to Kinetic Energy.'},
        {'sec_idx': 3, 'type': 'video', 'title': 'Roller Coaster Physics', 'text': 'Watch how energy transforms.', 'video_url': 'https://www.youtube.com/watch?v=Jnj8mc04r9E'},
        {'sec_idx': 4, 'type': 'concept_helper', 'title': 'Law of Conservation', 'text': 'Energy cannot be created or destroyed, only transformed.'},
        {'sec_idx': 4, 'type': 'flowchart', 'title': 'Energy Transformation', 'text': 'Visualizing how Potential Energy converts to Kinetic Energy.', 'image_url': 'https://mermaid.ink/img/pako:eNpVkMtqwzAQRX9FzKqF_IAeCqWbQsFQAqG7tciyxBZiS0ZSCyX_Xsdf4tJldTPn3DszGtToFCpoeD3pW_QeXwZ0h8-z_sQ12p05sB_tQ4B794dY6_tHj9F59GfW_4E-sB-sO9Z_sBfsC_vAPrAf7IA11v2wF-wL-8A-sB_s4J-x0k5bCg0ZylJyoOQYpZJMy5qrpRCSk0pWUp5S8oOQnJSkC_lLyU_2z7-Xw6GgUCqVbLhQ0pCpkHJYl0qJ4uO6Ff_2B2HqSgM?type=png'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Pendulums', 'text': 'A swinging pendulum constantly swaps PE and KE. It stops eventually only because of air resistance (friction).'}
    ],
    formulas=[
        {'text': 'W = F \\cdot d', 'label': 'Work', 'v1s': 'W', 'v1n': 'Work', 'v1u': 'J', 'v2s': 'F', 'v2n': 'Force', 'v2u': 'N', 'v3s': 'd', 'v3n': 'Distance', 'v3u': 'm'},
        {'text': 'KE = \\frac{1}{2}mv^2', 'label': 'Kinetic Energy', 'v1s': 'KE', 'v1n': 'Energy', 'v1u': 'J', 'v2s': 'm', 'v2n': 'Mass', 'v2u': 'kg', 'v3s': 'v', 'v3n': 'Velocity', 'v3u': 'm/s'},
        {'text': 'PE_g = mgh', 'label': 'Gravitational Potential Energy', 'v1s': 'PE', 'v1n': 'Potential Energy', 'v1u': 'J', 'v2s': 'm', 'v2n': 'Mass', 'v2u': 'kg', 'v3s': 'h', 'v3n': 'Height', 'v3u': 'm'}
    ],
    questions=[
        {'text': 'What is the unit for Work?', 'a': 'Newton', 'b': 'Watt', 'c': 'Joule', 'd': 'Meter', 'ans': 'C', 'exp': 'Work is measured in Joules (N·m).'},
        {'text': 'A ball held 2m high has what type of energy?', 'a': 'Kinetic', 'b': 'Gravitational Potential', 'c': 'Elastic', 'd': 'Thermal', 'ans': 'B', 'exp': 'It has potential due to gravity.'},
        {'text': 'If you lift a 5kg box 2 meters, how much work did you do? (g=10)', 'a': '10 J', 'b': '50 J', 'c': '100 J', 'd': '7 J', 'ans': 'C', 'exp': 'W = Fd = mgd = 5 * 10 * 2 = 100 J.'},
        {'text': 'Energy of motion is called...', 'a': 'Potential', 'b': 'Kinetic', 'c': 'Thermal', 'd': 'Chemical', 'ans': 'B', 'exp': 'Kinetic comes from the Greek word "kinesis" (motion).'},
        {'text': 'Can energy be destroyed?', 'a': 'Yes, by friction', 'b': 'Yes, in black holes', 'c': 'No, only transformed', 'd': 'No, except nuclear', 'ans': 'C', 'exp': 'Law of Conservation of Energy states it cannot be created or destroyed.'}
    ]
)

# 3. PHYSICS - Electricity
add_topic_data('phys-t3',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Circuits', 'icon': 'Zap', 'type': 'content'},
        {'title': "Ohm's Law", 'icon': 'Calculator', 'type': 'content'},
        {'title': 'Power & Safety', 'icon': 'AlertTriangle', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Understand circuit components', "Calculate using Ohm's Law", 'Differentiate series and parallel circuits', 'Calculate electrical power', 'Identify electrical safety hazards'],
    terms=[
        {'term': 'Voltage', 'def': 'Electrical potential difference (Volts)'},
        {'term': 'Current', 'def': 'Flow of electric charge (Amps)'},
        {'term': 'Resistance', 'def': 'Opposition to current flow (Ohms)'},
        {'term': 'Series Circuit', 'def': 'A circuit with only one path for current'},
        {'term': 'Parallel Circuit', 'def': 'A circuit with multiple paths for current'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Electric Circuits', 'text': 'A closed loop that allows current to flow. Requires a source (battery), load (bulb), and wires.', 'image_url': 'https://images.unsplash.com/photo-1549419163-e380e22784cb?w=800'},
        {'sec_idx': 2, 'type': 'concept_helper', 'title': 'Series vs Parallel', 'text': 'In Series, if one bulb goes out, they all go out. In Parallel, others stay on.'},
        {'sec_idx': 3, 'type': 'formula', 'title': "Ohm's Law", 'text': 'V = I \\cdot R'},
        {'sec_idx': 3, 'type': 'real_world', 'title': 'Resistors', 'text': 'Electronics use resistors to control current so delicate components don\'t burn out.'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Analogy', 'text': 'Voltage is like water pressure, Current is like water flow, Resistance is like a narrow pipe.'},
        {'sec_idx': 4, 'type': 'formula', 'title': 'Electrical Power', 'text': 'P = I \\cdot V'},
        {'sec_idx': 4, 'type': 'warning', 'title': 'Short Circuits', 'text': 'Never connect positive directly to negative without a load! It creates dangerous heat.'}
    ],
    formulas=[
        {'text': 'V = I \\cdot R', 'label': "Ohm's Law", 'v1s': 'V', 'v1n': 'Voltage', 'v1u': 'V', 'v2s': 'I', 'v2n': 'Current', 'v2u': 'A', 'v3s': 'R', 'v3n': 'Resistance', 'v3u': 'Ω'},
        {'text': 'P = I \\cdot V', 'label': "Electrical Power", 'v1s': 'P', 'v1n': 'Power', 'v1u': 'W', 'v2s': 'I', 'v2n': 'Current', 'v2u': 'A', 'v3s': 'V', 'v3n': 'Voltage', 'v3u': 'V'}
    ],
    questions=[
        {'text': 'What flows in a circuit?', 'a': 'Protons', 'b': 'Neutrons', 'c': 'Electrons', 'd': 'Atoms', 'ans': 'C', 'exp': 'Current is the flow of electrons.'},
        {'text': 'In which circuit type do all lights go out if one breaks?', 'a': 'Parallel', 'b': 'Series', 'c': 'Open', 'd': 'Short', 'ans': 'B', 'exp': 'Series circuits have only one path.'},
        {'text': 'If V=12V and R=4Ω, what is the Current?', 'a': '3 A', 'b': '48 A', 'c': '0.33 A', 'd': '16 A', 'ans': 'A', 'exp': 'I = V/R = 12/4 = 3 Amps.'},
        {'text': 'What unit measures Resistance?', 'a': 'Volt', 'b': 'Amp', 'c': 'Ohm', 'd': 'Watt', 'ans': 'C', 'exp': 'Ohms (Ω) measure resistance.'},
        {'text': 'What happens to current if resistance increases (Voltage constant)?', 'a': 'Increases', 'b': 'Decreases', 'c': 'Stays same', 'd': 'Becomes zero', 'ans': 'B', 'exp': 'Current and Resistance are inversely proportional.'}
    ]
)

# 4. MATH - Algebraic Expressions
add_topic_data('math-t1',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Basics', 'icon': 'BookOpen', 'type': 'content'},
        {'title': 'Simplifying', 'icon': 'Minimize2', 'type': 'content'},
        {'title': 'Expanding & Factoring', 'icon': 'Maximize2', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Identify variables and coefficients', 'Simplify like terms', 'Expand algebraic expressions using distributive property', 'Factor simple expressions', 'Evaluate expressions'],
    terms=[
        {'term': 'Variable', 'def': 'A letter representing an unknown number'},
        {'term': 'Coefficient', 'def': 'Number multiplying a variable'},
        {'term': 'Constant', 'def': 'A fixed value that does not change'},
        {'term': 'Like Terms', 'def': 'Terms that have identical variable parts'},
        {'term': 'Distributive Property', 'def': 'a(b + c) = ab + ac'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'text', 'title': 'What is Algebra?', 'text': 'Algebra is generalized arithmetic. We use letters to represent numbers we don\'t know yet.'},
        {'sec_idx': 2, 'type': 'image', 'title': 'Parts of an Expression', 'text': 'Visual breakdown of 3x + 5', 'image_url': 'https://upload.wikimedia.org/wikipedia/commons/thumb/1/15/Algebraic_term.svg/320px-Algebraic_term.svg.png'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Like Terms', 'text': 'You can only add terms if they have the same variable part. 2x + 3x = 5x, but 2x + 3y cannot be combined.'},
        {'sec_idx': 3, 'type': 'warning', 'title': 'Watch the powers', 'text': 'x and x² are NOT like terms!'},
        {'sec_idx': 4, 'type': 'formula', 'title': 'Distributive Property', 'text': 'a(b + c) = ab + ac'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Budgeting', 'text': 'If you buy 3 shirts for $x each and 2 pants for $y each, total cost is 3x + 2y.'},
        {'sec_idx': 4, 'type': 'text', 'title': 'Factoring', 'text': 'Factoring is the reverse of expanding. 2x + 4 = 2(x + 2).'}
    ],
    formulas=[
         {'text': 'a(b + c) = ab + ac', 'label': 'Distributive Property', 'v1s': 'a', 'v1n': 'Factor', 'v1u': '', 'v2s': 'b', 'v2n': 'Term 1', 'v2u': '', 'v3s': 'c', 'v3n': 'Term 2', 'v3u': ''}
    ],
    questions=[
        {'text': 'Simplify: 3x + 4y - x', 'a': '7xy', 'b': '2x + 4y', 'c': '6xy', 'd': '3x + 3y', 'ans': 'B', 'exp': 'Combine 3x and -x to get 2x. 4y stays separate.'},
        {'text': 'Expand: 2(x + 3)', 'a': '2x + 3', 'b': '2x + 6', 'c': 'x + 6', 'd': '5x', 'ans': 'B', 'exp': 'Multiply 2 by both terms inside: 2*x + 2*3.'},
        {'text': 'What is the coefficient in 5y?', 'a': 'y', 'b': '5', 'c': '5y', 'd': 'Unknown', 'ans': 'B', 'exp': 'The number multiplying the variable is the coefficient.'},
        {'text': 'Are 3x and 3x² like terms?', 'a': 'Yes', 'b': 'No', 'c': 'Sometimes', 'd': 'Only if x=1', 'ans': 'B', 'exp': 'No, because the exponents are different.'},
        {'text': 'Evaluate 2x + 1 when x = 4', 'a': '6', 'b': '7', 'c': '9', 'd': '8', 'ans': 'C', 'exp': '2(4) + 1 = 8 + 1 = 9.'}
    ]
)

# 5. MATH - Geometry: Triangles
add_topic_data('math-t2',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Types', 'icon': 'Triangle', 'type': 'content'},
        {'title': 'Pythagoras', 'icon': 'Calculator', 'type': 'content'},
        {'title': 'Area & Perimeter', 'icon': 'Grid', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Classify triangles by sides and angles', 'Use Pythagorean theorem', 'Calculate area of triangles', 'Identify triangle properties', 'Solve real-world problems involving triangles'],
    terms=[
        {'term': 'Hypotenuse', 'def': 'Longest side of a right triangle'},
        {'term': 'Isosceles', 'def': 'Triangle with 2 equal sides'},
        {'term': 'Equilateral', 'def': 'Triangle with 3 equal sides'},
        {'term': 'Scalene', 'def': 'Triangle with no equal sides'},
        {'term': 'Right Angle', 'def': '90 degree angle'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Triangle Types', 'text': 'Triangles can be classified by sides (equilateral, isosceles, scalene) or angles (acute, obtuse, right).', 'image_url': 'https://images.unsplash.com/photo-1616469829941-c7200ed5dabd?w=800'},
        {'sec_idx': 2, 'type': 'concept_helper', 'title': 'Angle Sum', 'text': 'The sum of angles in ANY triangle is always 180°.'},
        {'sec_idx': 3, 'type': 'formula', 'title': 'Pythagorean Theorem', 'text': 'a^2 + b^2 = c^2'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Usage', 'text': 'Used to find a missing side in a right-angled triangle. c is always the hypotenuse.'},
        {'sec_idx': 4, 'type': 'formula', 'title': 'Area of Triangle', 'text': 'A = \\frac{1}{2}bh'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Construction', 'text': 'Builders use the 3-4-5 rule (Pythagoras) to check if corners are perfectly square.'},
        {'sec_idx': 4, 'type': 'warning', 'title': 'Height must be perpendicular', 'text': 'When calculating area, the height must be at a 90° angle to the base.'}
    ],
    formulas=[
        {'text': 'a^2 + b^2 = c^2', 'label': 'Pythagorean Theorem', 'v1s': 'c', 'v1n': 'Hypotenuse', 'v1u': '', 'v2s': 'a', 'v2n': 'Side A', 'v2u': '', 'v3s': 'b', 'v3n': 'Side B', 'v3u': ''},
        {'text': 'A = \\frac{1}{2}b \\cdot h', 'label': 'Area of Triangle', 'v1s': 'A', 'v1n': 'Area', 'v1u': 'units²', 'v2s': 'b', 'v2n': 'Base', 'v2u': 'units', 'v3s': 'h', 'v3n': 'Height', 'v3u': 'units'}
    ],
    questions=[
        {'text': 'Which triangle has all equal sides?', 'a': 'Isosceles', 'b': 'Scalene', 'c': 'Equilateral', 'd': 'Right', 'ans': 'C', 'exp': 'Equi-lateral means equal sides.'},
        {'text': 'Calculate the hypotenuse if sides are 3 and 4.', 'a': '5', 'b': '6', 'c': '7', 'd': '25', 'ans': 'A', 'exp': '3² + 4² = 9 + 16 = 25. √25 = 5.'},
        {'text': 'Sum of angles in a triangle?', 'a': '90', 'b': '180', 'c': '360', 'd': '100', 'ans': 'B', 'exp': 'Always 180 degrees.'},
        {'text': 'Area of a triangle with base 10 and height 5?', 'a': '50', 'b': '25', 'c': '15', 'd': '100', 'ans': 'B', 'exp': '0.5 * 10 * 5 = 25.'},
        {'text': 'A triangle with angles 90, 45, 45 is...', 'a': 'Obtuse', 'b': 'Acute', 'c': 'Right Isosceles', 'd': 'Equilateral', 'ans': 'C', 'exp': 'Right (has 90°) and Isosceles (two equal angles/sides).'}
    ]
)

# 6. MATH - Probability
add_topic_data('math-t3',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Chance', 'icon': 'HelpCircle', 'type': 'content'},
        {'title': 'Calculating', 'icon': 'Calculator', 'type': 'content'},
        {'title': 'Multiple Events', 'icon': 'GitBranch', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Understand probability scale', 'Calculate simple probabilities', 'Determine sample spaces', 'Calculate probability of multiple events', 'Understand complementary events'],
    terms=[
        {'term': 'Event', 'def': 'An outcome or set of outcomes'},
        {'term': 'Sample Space', 'def': 'Set of all possible outcomes'},
        {'term': 'Impossible', 'def': 'Probability of 0'},
        {'term': 'Certain', 'def': 'Probability of 1 (or 100%)'},
        {'term': 'Independent Events', 'def': 'Events where one outcome does not affect the other'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'text', 'title': 'The Scale', 'text': 'Probability ranges from 0 (Impossible) to 1 (Certain). Fractions, decimals, or percentages can be used.'},
        {'sec_idx': 2, 'type': 'image', 'title': 'Probability Scale', 'text': '0 ----- 0.5 ----- 1', 'image_url': 'https://upload.wikimedia.org/wikipedia/commons/thumb/d/d7/Probability_Scale_Line.svg/640px-Probability_Scale_Line.svg.png'},
        {'sec_idx': 3, 'type': 'formula', 'title': 'Basic Probability', 'text': 'P(A) = \\frac{\\text{favorable outcomes}}{\\text{total outcomes}}'},
        {'sec_idx': 3, 'type': 'real_world', 'title': 'Dice', 'text': 'Rolling a 6 on a standard die has a 1/6 chance.', 'image_url': 'https://images.unsplash.com/photo-1595113316349-9fa4eb24f884?w=800'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Complementary Events', 'text': 'P(Not A) = 1 - P(A). Chance of rain is 20%, chance of NO rain is 80%.'},
        {'sec_idx': 4, 'type': 'text', 'title': 'Multiple Events', 'text': 'For independent events (like flipping two coins), multiply the probabilities.'},
        {'sec_idx': 4, 'type': 'warning', 'title': 'Gambler\'s Fallacy', 'text': 'Past results do not affect independent future results. The coin doesn\'t "remember" it was heads.'}
    ],
    formulas=[
        {'text': 'P(A) = \\frac{n(A)}{n(S)}', 'label': 'Probability', 'v1s': 'P', 'v1n': 'Probability', 'v1u': '', 'v2s': 'n(A)', 'v2n': 'Favorable', 'v2u': '', 'v3s': 'n(S)', 'v3n': 'Total', 'v3u': ''},
        {'text': 'P(A \\cap B) = P(A) \\times P(B)', 'label': 'Independent Events', 'v1s': 'P', 'v1n': 'Probability', 'v1u': '', 'v2s': 'A', 'v2n': 'Event A', 'v2u': '', 'v3s': 'B', 'v3n': 'Event B', 'v3u': ''}
    ],
    questions=[
        {'text': 'Probability of flipping heads?', 'a': '0.25', 'b': '0.5', 'c': '0.75', 'd': '1.0', 'ans': 'B', 'exp': '1 favorable (heads) / 2 total (heads, tails) = 0.5'},
        {'text': 'Probability of rolling a 7 on a standard die?', 'a': '1/6', 'b': '1/2', 'c': '0', 'd': '1', 'ans': 'C', 'exp': 'Impossible. Die only goes to 6.'},
        {'text': 'If P(Win) = 0.4, what is P(Lose)?', 'a': '0.4', 'b': '0.6', 'c': '0.5', 'd': '0.1', 'ans': 'B', 'exp': '1 - 0.4 = 0.6'},
        {'text': 'Probability of flipping heads TWICE in a row?', 'a': '0.5', 'b': '0.25', 'c': '0.1', 'd': '0.75', 'ans': 'B', 'exp': '0.5 * 0.5 = 0.25'},
        {'text': 'The set of all possible outcomes is called...', 'a': 'Event', 'b': 'Probability', 'c': 'Sample Space', 'd': 'Result', 'ans': 'C', 'exp': 'Definition of Sample Space.'}
    ]
)

# 7. CHEMISTRY - Atomic Structure
add_topic_data('chem-t1',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'The Atom', 'icon': 'Atom', 'type': 'content'},
        {'title': 'Subatomic Particles', 'icon': 'Disc', 'type': 'content'},
        {'title': 'Isotopes', 'icon': 'Copy', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Describe the structure of an atom', 'Identify protons, neutrons, electrons', 'Determine atomic mass and atomic number', 'Understand isotopes', 'Draw simple atomic models'],
    terms=[
        {'term': 'Nucleus', 'def': 'Central part of atom containing protons/neutrons'},
        {'term': 'Electron Shell', 'def': 'Region where electrons orbit'},
        {'term': 'Atomic Number', 'def': 'Number of protons (defines the element)'},
        {'term': 'Mass Number', 'def': 'Protons + Neutrons'},
        {'term': 'Isotope', 'def': 'Same element with different number of neutrons'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Building Blocks', 'text': 'All matter is made of atoms. They are the smallest unit of an element.', 'image_url': 'https://images.unsplash.com/photo-1614730341194-75c60740a070?w=800'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Inside the Atom', 'text': 'Protons (+ charge) and Neutrons (no charge) are in the center. Electrons (- charge) zoom around the outside.'},
        {'sec_idx': 3, 'type': 'video', 'title': 'Atomic Model', 'text': 'Visualizing the atom.', 'video_url': 'https://www.youtube.com/watch?v=IO9WS_HNmyg'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Empty Space', 'text': 'Atoms are mostly empty space. If the nucleus was a marble, the atom would be a stadium!'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Carbon Dating', 'text': 'We use Carbon-14 (an isotope) to figure out how old ancient fossils are.'},
        {'sec_idx': 4, 'type': 'image', 'title': 'Bohr Model', 'text': 'Simplified view of electron shells', 'image_url': 'https://upload.wikimedia.org/wikipedia/commons/thumb/5/55/Bohr-atom-PAR.svg/320px-Bohr-atom-PAR.svg.png'}
    ],
    formulas=[
        {'text': 'A = Z + N', 'label': 'Mass Number', 'v1s': 'A', 'v1n': 'Mass No', 'v1u': '', 'v2s': 'Z', 'v2n': 'Protons', 'v2u': '', 'v3s': 'N', 'v3n': 'Neutrons', 'v3u': ''}
    ],
    questions=[
        {'text': 'Which particle has a positive charge?', 'a': 'Electron', 'b': 'Neutron', 'c': 'Proton', 'd': 'Photon', 'ans': 'C', 'exp': 'Protons are positive (+)'},
        {'text': 'Where are electrons found?', 'a': 'Nucleus', 'b': 'Shells', 'c': 'Inside protons', 'd': 'Everywhere', 'ans': 'B', 'exp': 'Electrons orbit in shells/clouds around the nucleus.'},
        {'text': 'Atomic number tells you the number of...', 'a': 'Neutrons', 'b': 'Electrons', 'c': 'Protons', 'd': 'Isotopes', 'ans': 'C', 'exp': 'Atomic number = number of protons.'},
        {'text': 'Isotopes have different numbers of...', 'a': 'Protons', 'b': 'Neutrons', 'c': 'Electrons', 'd': 'Shells', 'ans': 'B', 'exp': 'Isotopes are same element (same protons) but different mass (neutrons).'},
        {'text': 'Most of an atom is...', 'a': 'Solid', 'b': 'Liquid', 'c': 'Empty space', 'd': 'Gas', 'ans': 'C', 'exp': 'The nucleus is tiny compared to the electron cloud volume.'}
    ]
)

# 8. CHEMISTRY - Periodic Table
add_topic_data('chem-t2',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Organization', 'icon': 'Grid', 'type': 'content'},
        {'title': 'Groups & Periods', 'icon': 'ArrowDown', 'type': 'content'},
        {'title': 'Metals vs Non-Metals', 'icon': 'Layers', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Read the Periodic Table', 'Understand Groups and Periods', 'Predict properties based on location', 'Identify metals, non-metals, and metalloids', 'Know common element families'],
    terms=[
        {'term': 'Group', 'def': 'Vertical column (similar properties)'},
        {'term': 'Period', 'def': 'Horizontal row (electron shells)'},
        {'term': 'Alkali Metals', 'def': 'Group 1 elements (highly reactive)'},
        {'term': 'Noble Gases', 'def': 'Group 18 elements (unreactive)'},
        {'term': 'Halogens', 'def': 'Group 17 elements (reactive non-metals)'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'The Map of Elements', 'text': 'The periodic table organizes all known elements by atomic number and chemical properties.', 'image_url': 'https://images.unsplash.com/photo-1603126857599-f6e157fa2fe6?w=800'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Navigation', 'text': 'Columns are called Groups (elements behave similarly). Rows are called Periods.'},
        {'sec_idx': 3, 'type': 'real_world', 'title': 'Noble Gases', 'text': 'Group 18 elements are "Noble Gases" - they are very stable and don\'t like to react (like Neon signs).'},
        {'sec_idx': 4, 'type': 'text', 'title': 'Metals vs Non-Metals', 'text': 'Metals are on the left (shiny, conduct), Non-metals on the right (dull, insulate). Staircase line separates them.'},
        {'sec_idx': 4, 'type': 'warning', 'title': 'Hydrogen Exception', 'text': 'Hydrogen is in Group 1 but it is a NON-METAL gas, not a metal.'},
        {'sec_idx': 4, 'type': 'image', 'title': 'Periodic Table Sections', 'text': 'Color coded regions', 'image_url': 'https://upload.wikimedia.org/wikipedia/commons/thumb/3/30/Periodic_Table_Structure.svg/640px-Periodic_Table_Structure.svg.png'}
    ],
    formulas=[],
    questions=[
        {'text': 'Elements in the same column usually have...', 'a': 'Same mass', 'b': 'Similar properties', 'c': 'Same atomic number', 'd': 'Different states', 'ans': 'B', 'exp': 'Groups (columns) share chemical properties.'},
        {'text': 'Which group contains the Noble Gases?', 'a': '1', 'b': '2', 'c': '17', 'd': '18', 'ans': 'D', 'exp': 'Group 18 (far right) are Noble Gases.'},
        {'text': 'Where are metals found on the table?', 'a': 'Left', 'b': 'Right', 'c': 'Top only', 'd': 'Bottom only', 'ans': 'A', 'exp': 'Metals make up the majority of the left side.'},
        {'text': 'Horizontal rows are called...', 'a': 'Groups', 'b': 'Families', 'c': 'Periods', 'd': 'Sections', 'ans': 'C', 'exp': 'Periods go across.'},
        {'text': 'Which element is a non-metal in Group 1?', 'a': 'Lithium', 'b': 'Sodium', 'c': 'Hydrogen', 'd': 'Potassium', 'ans': 'C', 'exp': 'Hydrogen is the exception.'}
    ]
)

# 9. CHEMISTRY - Bonding
add_topic_data('chem-t3',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Why Bond?', 'icon': 'Link', 'type': 'content'},
        {'title': 'Ionic Bonding', 'icon': 'Zap', 'type': 'content'},
        {'title': 'Covalent Bonding', 'icon': 'GitCommit', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Explain why atoms bond', 'Distinguish ionic and covalent bonds', 'Draw dot and cross diagrams', 'Predict bond type between elements', 'Name simple compounds'],
    terms=[
        {'term': 'Ion', 'def': 'Atom with a charge (lost or gained electrons)'},
        {'term': 'Molecule', 'def': 'Group of atoms bonded together'},
        {'term': 'Ionic Bond', 'def': 'Transfer of electrons (Metal + Non-Metal)'},
        {'term': 'Covalent Bond', 'def': 'Sharing of electrons (Non-Metal + Non-Metal)'},
        {'term': 'Valence Shell', 'def': 'Outermost electron shell'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'text', 'title': 'Stability', 'text': 'Atoms bond to become stable, usually by getting a full outer shell of electrons (Octet Rule).'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Ionic Bonding', 'text': 'One atom STEALS electrons from another. Creates + and - ions that attract.'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Metal + Non-Metal', 'text': 'Ionic bonds usually happen between a metal and a non-metal (e.g., NaCl).'},
        {'sec_idx': 4, 'type': 'text', 'title': 'Covalent Bonding', 'text': 'Atoms SHARE electrons. Like two people holding hands.'},
        {'sec_idx': 4, 'type': 'image', 'title': 'Water Molecule', 'text': 'H2O is a covalent bond', 'image_url': 'https://images.unsplash.com/photo-1532634993-15f421e42ec0?w=800'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Salt vs Sugar', 'text': 'Salt is Ionic (high melting point), Sugar is Covalent (low melting point).'}
    ],
    formulas=[],
    questions=[
        {'text': 'In a covalent bond, electrons are...', 'a': 'Transferred', 'b': 'Destroyed', 'c': 'Shared', 'd': 'Doubled', 'ans': 'C', 'exp': 'Co-valent means sharing valence electrons.'},
        {'text': 'Ionic bonds occur between...', 'a': 'Two metals', 'b': 'Two non-metals', 'c': 'Metal and Non-metal', 'd': 'Noble gases', 'ans': 'C', 'exp': 'Opposites attract (Metal loses, Non-metal gains).'},
        {'text': 'What charge does an atom get if it loses an electron?', 'a': 'Positive', 'b': 'Negative', 'c': 'Neutral', 'd': 'Unknown', 'ans': 'A', 'exp': 'Losing a negative electron leaves a net positive charge.'},
        {'text': 'NaCl (Table Salt) is...', 'a': 'Covalent', 'b': 'Ionic', 'c': 'Metallic', 'd': 'Magnetic', 'ans': 'B', 'exp': 'Sodium (Metal) + Chlorine (Non-Metal).'},
        {'text': 'Why do atoms bond?', 'a': 'To get bigger', 'b': 'To become unstable', 'c': 'To fill outer shell', 'd': 'To change element', 'ans': 'C', 'exp': 'Full outer shell means stability.'}
    ]
)

# 10. BIOLOGY - Cell Structure
add_topic_data('bio-t1',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Cell Theory', 'icon': 'BookOpen', 'type': 'content'},
        {'title': 'Organelles', 'icon': 'Circle', 'type': 'content'},
        {'title': 'Plant vs Animal', 'icon': 'Leaf', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['State Cell Theory', 'Identify function of nucleus, mitochondria, cell membrane', 'Distinguish plant and animal cells', 'Understand specialized cells', 'Explain diffusion'],
    terms=[
        {'term': 'Organelle', 'def': 'Specialized structure within a cell'},
        {'term': 'Prokaryote', 'def': 'Simple cell without nucleus (bacteria)'},
        {'term': 'Eukaryote', 'def': 'Complex cell with nucleus (plants, animals)'},
        {'term': 'Chloroplast', 'def': 'Site of photosynthesis in plants'},
        {'term': 'Cell Wall', 'def': 'Rigid outer layer of plant cells'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Unit of Life', 'text': 'Cells are the basic structural and functional units of life.', 'image_url': 'https://images.unsplash.com/photo-1530210124550-912dc1381cb8?w=800'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Mitochondria', 'text': 'The POWERHOUSE of the cell. Generates energy (ATP).'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Nucleus', 'text': 'The BRAIN. Contains DNA and controls cell activity.'},
        {'sec_idx': 4, 'type': 'concept_helper', 'title': 'Plant Differences', 'text': 'Plant cells have Cell Walls and Chloroplasts. Animal cells do not.'},
        {'sec_idx': 4, 'type': 'flowchart', 'title': 'Animal vs Plant', 'text': 'Comparison diagram', 'image_url': 'https://mermaid.ink/img/pako:eNpVkEFqwzAQRf8iZtVC_IAeCqWbQsFQAqG7tciyxBZiS0ZSCyX_Xsdf4tJldfPnzZtRo1OooOH1pG_Re3wZ0B0-z_oT12h35sB-tA8B7t0fYq3vHz1G59GfWf8H-sB-sO5Y_8FesC_sA_vAfrAD1lj3w16wL-wD-8B-sIN_xko7bSk0ZChLyYGSY5RKMitrLpdCSM4qWUv5UsmvQnJWkq7kLyU_2T__Xg6HgkKlVLLhQklD5kLKcV0pJYqP61b82x-HqSgN?type=png'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Specialized Cells', 'text': 'Red blood cells have no nucleus to carry more oxygen. Nerve cells are long to send signals.'}
    ],
    formulas=[],
    questions=[
        {'text': 'Which organelle produces energy?', 'a': 'Ribosome', 'b': 'Nucleus', 'c': 'Mitochondria', 'd': 'Vacuole', 'ans': 'C', 'exp': 'Mitochondria perform cellular respiration to make ATP.'},
        {'text': 'What is found in plant cells but NOT animal cells?', 'a': 'Nucleus', 'b': 'Cell Wall', 'c': 'Mitochondria', 'd': 'Cell Membrane', 'ans': 'B', 'exp': 'Cell Wall provides rigid structure for plants.'},
        {'text': 'Control center of the cell?', 'a': 'Nucleus', 'b': 'Cytoplasm', 'c': 'Membrane', 'd': 'Golgi', 'ans': 'A', 'exp': 'Nucleus holds DNA instructions.'},
        {'text': 'Simple cells like bacteria are...', 'a': 'Eukaryotes', 'b': 'Prokaryotes', 'c': 'Animals', 'd': 'Plants', 'ans': 'B', 'exp': 'Prokaryotes lack a membrane-bound nucleus.'},
        {'text': 'Photosynthesis happens in...', 'a': 'Mitochondria', 'b': 'Chloroplasts', 'c': 'Ribosomes', 'd': 'Vacuoles', 'ans': 'B', 'exp': 'Chloroplasts contain chlorophyll for photosynthesis.'}
    ]
)

# 11. BIOLOGY - Genetics
add_topic_data('bio-t2',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'DNA', 'icon': 'Dna', 'type': 'content'},
        {'title': 'Heredity', 'icon': 'GitBranch', 'type': 'content'},
        {'title': 'Punnett Squares', 'icon': 'Grid', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Describe DNA structure', 'Understand basic inheritance', 'Use Punnett Squares', 'Define genotype and phenotype', 'Understand mutations'],
    terms=[
        {'term': 'Gene', 'def': 'Unit of heredity'},
        {'term': 'Allele', 'def': 'Variant form of a gene (e.g., Blue vs Brown eyes)'},
        {'term': 'Dominant', 'def': 'Trait that shows up if present (Capital letter)'},
        {'term': 'Recessive', 'def': 'Trait that is hidden by dominant (Lowercase)'},
        {'term': 'Genotype', 'def': 'Genetic makeup (e.g., Bb)'},
        {'term': 'Phenotype', 'def': 'Physical appearance (e.g., Brown eyes)'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Blueprint of Life', 'text': 'DNA holds the instructions for building and operating an organism.', 'image_url': 'https://images.unsplash.com/photo-1576086213369-97a306d36557?w=800'},
        {'sec_idx': 2, 'type': 'text', 'title': 'Double Helix', 'text': 'DNA looks like a twisted ladder. The rungs are base pairs (A-T, C-G).'},
        {'sec_idx': 3, 'type': 'concept_helper', 'title': 'Dominant vs Recessive', 'text': 'Dominant traits (like brown eyes) often hide recessive traits (like blue eyes).'},
        {'sec_idx': 4, 'type': 'text', 'title': 'Punnett Squares', 'text': 'A tool to predict the probability of offspring traits.'},
        {'sec_idx': 4, 'type': 'image', 'title': 'Punnett Square Example', 'text': 'Crossing Bb x Bb', 'image_url': 'https://upload.wikimedia.org/wikipedia/commons/thumb/2/22/Punnett_Square.svg/320px-Punnett_Square.svg.png'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Inheritance', 'text': 'You get half your DNA from mom and half from dad.'}
    ],
    formulas=[],
    questions=[
        {'text': 'What molecule carries genetic info?', 'a': 'Protein', 'b': 'Carbohydrate', 'c': 'DNA', 'd': 'Lipid', 'ans': 'C', 'exp': 'Deoxyribonucleic Acid.'},
        {'text': 'Shape of DNA?', 'a': 'Single Helix', 'b': 'Double Helix', 'c': 'Circle', 'd': 'Square', 'ans': 'B', 'exp': 'Twisted ladder shape.'},
        {'text': 'If B is Brown (dominant) and b is blue (recessive), what is Bb?', 'a': 'Blue', 'b': 'Brown', 'c': 'Green', 'd': 'Mix', 'ans': 'B', 'exp': 'Dominant B masks recessive b.'},
        {'text': 'Physical appearance is called...', 'a': 'Genotype', 'b': 'Phenotype', 'c': 'Karyotype', 'd': 'Biotype', 'ans': 'B', 'exp': 'Pheno = Physical.'},
        {'text': 'Probability of bb from Bb x Bb?', 'a': '0%', 'b': '25%', 'c': '50%', 'd': '100%', 'ans': 'B', 'exp': '1 out of 4 squares will be bb.'}
    ]
)

# 12. BIOLOGY - Ecosystems
add_topic_data('bio-t3',
    sections=[
        {'title': 'Objectives', 'icon': 'Target', 'type': 'objectives'},
        {'title': 'Components', 'icon': 'Globe', 'type': 'content'},
        {'title': 'Food Webs', 'icon': 'Share2', 'type': 'content'},
        {'title': 'Cycles', 'icon': 'RefreshCw', 'type': 'content'},
        {'title': 'Quiz', 'icon': 'HelpCircle', 'type': 'quiz'}
    ],
    objectives=['Define ecosystem', 'Trace energy flow in food webs', 'Distinguish biotic and abiotic factors', 'Understand carbon and water cycles', 'Identify trophic levels'],
    terms=[
        {'term': 'Biotic', 'def': 'Living components (plants, animals)'},
        {'term': 'Abiotic', 'def': 'Non-living components (sun, water, soil)'},
        {'term': 'Producer', 'def': 'Organism that makes its own food (plants)'},
        {'term': 'Consumer', 'def': 'Organism that eats others'},
        {'term': 'Decomposer', 'def': 'Breaks down dead matter'}
    ],
    content=[
        {'sec_idx': 2, 'type': 'introduction', 'title': 'Web of Life', 'text': 'An ecosystem includes all living things in an area interacting with each other and their environment.', 'image_url': 'https://images.unsplash.com/photo-1470071459604-3b5ec3a7fe05?w=800'},
        {'sec_idx': 2, 'type': 'concept_helper', 'title': 'Biotic vs Abiotic', 'text': 'Biotic = Living (Wolf). Abiotic = Non-living (Water).'},
        {'sec_idx': 3, 'type': 'text', 'title': 'Energy Flow', 'text': 'Energy starts from the Sun -> Producers -> Consumers.'},
        {'sec_idx': 3, 'type': 'warning', 'title': 'Energy Loss', 'text': '90% of energy is lost at each level (heat, movement). Only 10% is passed on.'},
        {'sec_idx': 3, 'type': 'flowchart', 'title': 'Food Chain', 'text': 'Sun -> Grass -> Rabbit -> Fox', 'image_url': 'https://mermaid.ink/img/pako:eNpVkM1qwzAQhF9FzKqF9AN6KJRQKBhKIHS3FllWbCG2ZCSt0JL3Xsdf4tKldTPfzGhGo1OooOH1rO_RB3wd0B0-z_oT12h35sB-tA8B7t0fYu3vHz3G4NGfWf8H-sB-sO5Y_8FesC_sA_vAfrAD1lj3w16wL-wD-8B-sIN_xko7bSk0ZChLyYGSY5RKMitrLpdCSM4qWUv5UsmvQnJWkq7kLyU_2T__Xg6HgkKlVLLhQklD5kLKcV0pJYqP61b82x-HqSgN?type=png'},
        {'sec_idx': 4, 'type': 'real_world', 'title': 'Decomposers', 'text': 'Fungi and bacteria recycle nutrients back into the soil. Without them, we would be buried in waste!'}
    ],
    formulas=[],
    questions=[
        {'text': 'Which is an abiotic factor?', 'a': 'Tree', 'b': 'Bacteria', 'c': 'Sunlight', 'd': 'Wolf', 'ans': 'C', 'exp': 'Sunlight is non-living.'},
        {'text': 'Organisms that make their own food are...', 'a': 'Consumers', 'b': 'Producers', 'c': 'Decomposers', 'd': 'Predators', 'ans': 'B', 'exp': 'Producers (plants) use photosynthesis.'},
        {'text': 'How much energy is passed to the next level?', 'a': '100%', 'b': '50%', 'c': '10%', 'd': '0%', 'ans': 'C', 'exp': '10% rule. Rest is lost as heat.'},
        {'text': 'A network of interconnected food chains is a...', 'a': 'Food Web', 'b': 'Pyramid', 'c': 'Cycle', 'd': 'Biome', 'ans': 'A', 'exp': 'Web implies multiple connections.'},
        {'text': 'What is the primary source of energy for Earth?', 'a': 'The Moon', 'b': 'Volcanoes', 'c': 'The Sun', 'd': 'Wind', 'ans': 'C', 'exp': 'Sunlight drives photosynthesis.'}
    ]
)


# ============================================================================
# FUNCTIONS
# ============================================================================

def create_sample_excel(output_path='public/StudyHub_Complete_Data.xlsx'):
    """Create a sample Excel file with all the required sheets and data."""
    print(f"Creating sample Excel file: {output_path}")
    
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for i, (sheet_name, data) in enumerate(SAMPLE_DATA.items()):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        
        schema = SHEET_SCHEMAS.get(sheet_name, {})
        columns = schema.get('columns', list(data[0].keys()) if data else [])
        
        # Write headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = max(len(str(col_name)), max(len(str(row.get(col_name, ''))) for row in data) if data else 0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    print(f"✅ Sample Excel file created: {output_path}")
    print(f"   Sheets created: {', '.join(SAMPLE_DATA.keys())}")
    return output_path


def validate_excel(file_path):
    """Validate an Excel file against the required schema."""
    print(f"Validating: {file_path}")
    errors = []
    warnings = []
    
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"❌ Error opening file: {e}")
        return False
    
    # Check for required sheets
    for sheet_name, schema in SHEET_SCHEMAS.items():
        if sheet_name not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet_name}")
            continue
        
        ws = wb[sheet_name]
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1] if cell.value]
        headers_lower = [h.lower().replace(' ', '_') if h else '' for h in headers]
        
        # Check required columns
        for required_col in schema.get('required', []):
            if required_col not in headers_lower:
                errors.append(f"{sheet_name}: Missing required column '{required_col}'")
        
        # Check for data
        if ws.max_row < 2:
            warnings.append(f"{sheet_name}: No data rows found")
        
        # Validate content types for Study_Content
        if sheet_name == 'Study_Content' and 'content_type' in headers_lower:
            type_col = headers_lower.index('content_type') + 1
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=type_col).value
                if cell_value and cell_value not in CONTENT_TYPES:
                    warnings.append(f"{sheet_name} row {row}: Invalid content_type '{cell_value}'. Valid: {CONTENT_TYPES}")
        
        # Validate icons
        if sheet_name in ['Subjects', 'Topic_Sections', 'Achievements']:
            icon_col_name = 'icon' if sheet_name != 'Topic_Sections' else 'section_icon'
            if icon_col_name in headers_lower:
                icon_col = headers_lower.index(icon_col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=icon_col).value
                    if cell_value and cell_value not in VALID_ICONS:
                        warnings.append(f"{sheet_name} row {row}: Unknown icon '{cell_value}'")
    
    # Print results
    print("\n" + "="*50)
    if errors:
        print("❌ ERRORS:")
        for error in errors:
            print(f"   • {error}")
    
    if warnings:
        print("⚠️  WARNINGS:")
        for warning in warnings:
            print(f"   • {warning}")
    
    if not errors and not warnings:
        print("✅ Validation passed! No issues found.")
    elif not errors:
        print("✅ Validation passed with warnings.")
    else:
        print("❌ Validation failed. Please fix errors before using.")
    
    print("="*50)
    return len(errors) == 0


def export_to_json(file_path, output_path=None):
    """Export Excel data to JSON format for use without Google Sheets."""
    print(f"Exporting to JSON: {file_path}")
    
    if output_path is None:
        output_path = file_path.replace('.xlsx', '.json')
    
    wb = load_workbook(file_path)
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers
        headers = [cell.value.lower().replace(' ', '_') if cell.value else f'col_{i}' 
                   for i, cell in enumerate(ws[1])]
        
        # Get data
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value if value is not None else ''
                rows.append(row_dict)
        
        data[sheet_name] = rows
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"✅ JSON exported: {output_path}")
    return output_path


def print_schema():
    """Print the data schema for reference."""
    print("\n" + "="*60)
    print("STUDYHUB DATA SCHEMA")
    print("="*60)
    
    for sheet_name, schema in SHEET_SCHEMAS.items():
        print(f"\n📄 {sheet_name}")
        print(f"   Description: {schema['description']}")
        print(f"   Columns: {', '.join(schema['columns'])}")
        print(f"   Required: {', '.join(schema['required'])}")
    
    print("\n" + "-"*60)
    print("VALID CONTENT TYPES:", ', '.join(CONTENT_TYPES))
    print("VALID SECTION TYPES:", ', '.join(SECTION_TYPES))
    print("VALID ICONS:", ', '.join(VALID_ICONS))
    print("="*60)


def validate_coverage(file_path):
    """
    Validate that the data meets minimum content coverage requirements.
    - Each subject has >= 3 topics
    - Each topic has Learning Objectives
    - Each topic has Key Terms
    - Each topic has sufficient Handout Content (concept_helper, real_world, etc.)
    - Each topic has >= 3 Quiz Questions
    """
    print(f"Validating Content Coverage: {file_path}")

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"❌ Error opening file: {e}")
        return False

    def get_data(sheet_name):
        if sheet_name not in wb.sheetnames: return []
        ws = wb[sheet_name]
        headers = [c.value.lower().replace(' ', '_') if c.value else '' for c in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                data.append({k: v for k, v in zip(headers, row)})
        return data

    subjects = get_data('Subjects')
    topics = get_data('Topics')
    objectives = get_data('Learning_Objectives')
    terms = get_data('Key_Terms')
    content = get_data('Study_Content')
    questions = get_data('Quiz_Questions')

    errors = []

    # 1. Subject Coverage
    if len(subjects) < 4:
        errors.append(f"Expected at least 4 subjects, found {len(subjects)}")

    # 2. Topic Coverage per Subject
    for sub in subjects:
        sub_key = sub.get('subject_key')
        sub_topics = [t for t in topics if t.get('subject_key') == sub_key]
        if len(sub_topics) < 3:
            errors.append(f"Subject '{sub.get('name')}' has only {len(sub_topics)} topics (min 3 required)")

        # 3. Topic Content Coverage
        for topic in sub_topics:
            tid = topic.get('topic_id')
            tname = topic.get('topic_name')

            # Check Objectives
            t_objs = [o for o in objectives if o.get('topic_id') == tid]
            if not t_objs:
                errors.append(f"Topic '{tname}' ({tid}) missing Learning Objectives")

            # Check Terms
            t_terms = [t for t in terms if t.get('topic_id') == tid]
            if not t_terms:
                errors.append(f"Topic '{tname}' ({tid}) missing Key Terms")

            # Check Quiz
            t_quiz = [q for q in questions if q.get('topic_id') == tid]
            if len(t_quiz) < 3:
                errors.append(f"Topic '{tname}' ({tid}) has {len(t_quiz)} quiz questions (min 3 required)")

            # Check Handout Content
            # Valid types for handout: 'formula', 'concept_helper', 'warning', 'real_world', 'flowchart', 'image'
            valid_types = ['formula', 'concept_helper', 'warning', 'real_world', 'flowchart', 'image']
            t_content = [c for c in content if c.get('section_id', '').startswith(tid) and c.get('content_type') in valid_types]

            if not t_content:
                errors.append(f"Topic '{tname}' ({tid}) missing Handout-compatible content (concept_helper, real_world, etc.)")

    print("\n" + "="*50)
    if errors:
        print("❌ COVERAGE FAILED:")
        for e in errors:
            print(f"   • {e}")
        return False
    else:
        print("✅ Coverage Check Passed! All topics have required handouts, quizzes, and objectives.")
        return True


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description='StudyHub Data Setup Script')
    parser.add_argument('command', choices=['create-sample', 'validate', 'validate-coverage', 'export-json', 'schema'],
                       help='Command to run')
    parser.add_argument('file', nargs='?', help='Input file path (for validate/export-json)')
    parser.add_argument('-o', '--output', help='Output file path')
    
    args = parser.parse_args()
    
    if args.command == 'create-sample':
        output = args.output or 'public/StudyHub_Complete_Data.xlsx'
        create_sample_excel(output)
    
    elif args.command == 'validate':
        if not args.file:
            print("Error: Please provide a file path to validate")
            sys.exit(1)
        success = validate_excel(args.file)
        sys.exit(0 if success else 1)

    elif args.command == 'validate-coverage':
        if not args.file:
            print("Error: Please provide a file path to validate")
            sys.exit(1)
        success = validate_coverage(args.file)
        sys.exit(0 if success else 1)
    
    elif args.command == 'export-json':
        if not args.file:
            print("Error: Please provide a file path to export")
            sys.exit(1)
        export_to_json(args.file, args.output)
    
    elif args.command == 'schema':
        print_schema()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nQuick start:")
        print("  python setup_data.py create-sample   # Create sample Excel file")
        print("  python setup_data.py schema          # Show data schema")
    else:
        main()
